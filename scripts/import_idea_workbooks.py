#!/usr/bin/env python3
"""Normalize the PB11 and PB12 idea workbooks into the site's CSV schema."""

import csv
from collections import Counter
from pathlib import Path

from ftfy import fix_text
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "source-data"

PB11_WORKBOOK = SRC / "pb11_ideas.xlsx"
PB12_WORKBOOK = SRC / "pb12_ideas.xlsx"

OUTPUT_COLUMNS = [
    "PB Cycle",
    "Date Range",
    "Idea #",
    "Committee",
    "Project Title",
    "Project Description",
    "Project Status",
    "Idea Submitter",
    "Location",
    "Latitude",
    "Longitude",
    "Winning Project ID",
    "Idea Status",
    "Link to Other Information",
]


def text(value):
    if value is None:
        return ""
    return fix_text(str(value)).replace("_x000D_", "\n").strip()


def idea_ref(value):
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return text(value)


def sheet_rows(workbook, sheet_name):
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [text(value) for value in next(rows)]
    for values in rows:
        if any(value is not None for value in values):
            yield dict(zip(headers, values))


def normalize_pb11():
    workbook = load_workbook(PB11_WORKBOOK, read_only=True, data_only=True)
    normalized = []
    for sheet_name in ("CR", "Environment", "FPR", "TSS", "Youth"):
        for row in sheet_rows(workbook, sheet_name):
            idea_status = row.get("Existing Project Status") or row.get("Exisiting Project Status")
            location = text(row.get("Location"))
            if not location and row.get("Citywide") is True:
                location = "Citywide"
            normalized.append({
                "PB Cycle": "11",
                "Date Range": "August 2024-March 2025",
                "Idea #": idea_ref(row.get("Idea #")),
                "Committee": text(row.get("Committee")),
                "Project Title": text(row.get("Project Title")),
                "Project Description": text(row.get("Project Description")),
                "Project Status": text(row.get("Project Status")),
                "Idea Submitter": text(row.get("Idea Submitter")),
                "Location": location,
                "Latitude": text(row.get("Latitude")),
                "Longitude": text(row.get("Longitude")),
                "Winning Project ID": "",
                "Idea Status": text(idea_status),
                "Link to Other Information": "",
            })
    return normalized


# Corrections applied on top of the Budget Office PB12 workbook, reported to
# Melissa on 2026-09-03. Remove once fixed at the source.
PB12_ID_FIXES = {
    # "School Supplies Campaign" has the status text in its ID cell; 734090 is
    # the missing value in the surrounding submission-ID sequence.
    ("Advanced to ballot", "School Supplies Campaign"): "734090",
}
PB12_TEXT_FIXES = {
    ("732913", "Idea Status"): (
        "Inspird a winning ballot project!",
        "Inspired a winning ballot project!",
    ),
    ("732913", "Project Status"): (
        "A proposal for concrete barriers for bike lanes made it to the ballot "
        "nad won funding through the vote!",
        "A proposal for concrete barriers for bike lanes made it to the ballot "
        "and won funding through the vote!",
    ),
}


def normalize_pb12():
    workbook = load_workbook(PB12_WORKBOOK, read_only=True, data_only=True)
    normalized = []
    for row in sheet_rows(workbook, "PB12 Final"):
        ref = idea_ref(row.get("Idea #"))
        title = text(row.get("Project Title"))
        ref = PB12_ID_FIXES.get((ref, title), ref)
        record = {
            "PB Cycle": "12",
            "Date Range": "September 2025-March 2026",
            "Idea #": ref,
            "Committee": text(row.get("Committee")),
            "Project Title": title,
            "Project Description": text(row.get("Project Description")),
            "Project Status": text(row.get("Project Status Description")),
            "Idea Submitter": text(row.get("Idea Submitter")),
            "Location": text(row.get("Location")),
            "Latitude": text(row.get("Latitude")),
            "Longitude": text(row.get("Longitude")),
            "Winning Project ID": text(row.get("Winning Project ID")),
            "Idea Status": text(row.get("Idea Status")),
            "Link to Other Information": text(row.get("Link to Other Information")),
        }
        for (fix_ref, field), (old, new) in PB12_TEXT_FIXES.items():
            if ref == fix_ref and record[field] == old:
                record[field] = new
        normalized.append(record)
    return normalized


def write_csv(name, rows):
    output = SRC / name
    with output.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    refs = Counter(row["Idea #"] for row in rows)
    duplicates = sum(count - 1 for ref, count in refs.items() if ref and count > 1)
    nonnumeric = [ref for ref in refs if ref and not ref.replace(".", "", 1).isdigit()]
    print(
        f"{output.name}: {len(rows)} rows "
        f"({duplicates} repeated idea IDs, {len(nonnumeric)} nonnumeric idea IDs)"
    )


def main():
    write_csv("pb11_ideas.csv", normalize_pb11())
    write_csv("pb12_ideas.csv", normalize_pb12())


if __name__ == "__main__":
    main()
